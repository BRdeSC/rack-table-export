import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

#Função para estilizar a tabela
def apply_excel_styles(ws, title_text, headers, data, logo_path=None):
    logo_height_px = 70
    logo_width_px = 60
    
    # Inserir a logo
    if logo_path:
        img = Image(logo_path)
        img.height = logo_height_px
        img.width = logo_width_px
        ws.add_image(img, 'A1')

    # Definir a altura da linha 1 para acomodar a logo
    ws.row_dimensions[1].height = logo_height_px * 0.85

    # Título principal dinâmico
    num_cols = len(headers)
    end_col_letter = get_column_letter(num_cols)
    ws.merge_cells(f'A1:{end_col_letter}1')
    title_cell = ws['A1']
    title_cell.value = title_text
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Espaçamento
    #ws.append([])
    
    # Cabeçalho da tabela
    ws.append(headers)
    
    header_fill = PatternFill(start_color="0077BA", end_color="0077BA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    header_row_index = ws.max_row
    for cell in ws[header_row_index]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    # Adicionar dados à tabela
    for row in data:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
    
    # Ajustar largura das colunas com base no conteúdo
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width