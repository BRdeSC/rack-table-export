from flask import Blueprint, Response, jsonify, request
import MySQLdb
from src.utils.database import db_connection
from src.utils.excel_utils import apply_excel_styles
from src.utils.object_types import get_object_type_name 
from io import BytesIO
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

exports_bp = Blueprint('exports', __name__)

# Export Lista de Racks 
@exports_bp.route("/api/racks/export/xlsx")
@db_connection
def export_racks_xlsx(db):
    cursor = db.cursor()
    cursor.execute("""
        SELECT 
            r.name, r.location_name, r.row_name, 
            r.height,
            COUNT(DISTINCT rs.object_id) as object_count
        FROM Rack r
        LEFT JOIN RackSpace rs ON r.id = rs.rack_id
        GROUP BY r.id
        ORDER BY r.name
    """)
    data = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Racks"

    title = "Relatório de Racks"
    logo_path = "logo-coids.png" 
    headers = ["Rack", "Localizacao", "Linha", "Altura", "Equipamentos"]
    apply_excel_styles(ws, title, headers, data, logo_path)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=racks.xlsx"}
    )


# Export Lista de Equipamentos por Rack
@exports_bp.route("/api/racks/<int:rack_id>/export/xlsx")
@db_connection
def export_rack_equipment_xlsx(db, rack_id):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    
    # Buscar informações do rack
    cursor.execute("SELECT name, location_name, row_name, height FROM Rack WHERE id = %s", (rack_id,))
    rack_info = cursor.fetchone()
    
    if not rack_info:
        return jsonify({"error": "Rack não encontrado"}), 404
    
    # Buscar equipamentos do rack
    cursor.execute("""
        SELECT 
            o.id,
            o.name,
            o.objtype_id,
            o.asset_no,
            MIN(rs.unit_no) as start_unit,
            MAX(rs.unit_no) as end_unit,
            COUNT(DISTINCT rs.unit_no) as slot_count
        FROM RackSpace rs
        JOIN Object o ON rs.object_id = o.id
        WHERE rs.rack_id = %s
        GROUP BY o.id
        ORDER BY MIN(rs.unit_no) DESC
    """, (rack_id,))
    
    data = cursor.fetchall()

    # Processar dados
    processed_data = []
    for row in data:
        objtype_name = get_object_type_name(row['objtype_id'])
        
        # Calcular range de slots
        if row['slot_count'] > 1:
            slot_range = f"{row['start_unit']}-{row['end_unit']}U"
        else:
            slot_range = f"{row['start_unit']}U"
        
        processed_row = (
            slot_range,
            row['name'],
            objtype_name,
            row['asset_no'] or 'N/A',
            row['slot_count']
        )
        processed_data.append(processed_row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    
    title = f"Equipamentos do Rack: {rack_info['name']}"
    logo_path = "logo-coids.png"
    headers = ["Slots", "Nome", "Tipo", "Asset No.", "Altura (U)"]
    apply_excel_styles(ws, title, headers, processed_data, logo_path)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"rack_{rack_info['name'].lower().replace(' ', '_')}_equipamentos.xlsx"
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )


# Export Lista de Equipamentos 
@exports_bp.route("/api/objects/export/xlsx")
@db_connection
def export_all_objects_xlsx(db):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)  
    cursor.execute("""
        SELECT 
            o.id,
            o.name, 
            GROUP_CONCAT(DISTINCT l.name) as location_names,
            GROUP_CONCAT(DISTINCT r.name) as rack_names,
            o.objtype_id, 
            o.asset_no
        FROM Object o
        LEFT JOIN RackSpace rs ON o.id = rs.object_id
        LEFT JOIN Rack r ON rs.rack_id = r.id
        LEFT JOIN Location l ON r.location_id = l.id
        WHERE o.objtype_id NOT IN (1560, 1561, 1562)
        GROUP BY o.id
        ORDER BY o.name
    """)
    data = cursor.fetchall()

    # Processar dados para incluir objtype_name
    processed_data = []
    for row in data:
        objtype_name = get_object_type_name(row['objtype_id'])
        processed_row = (
            row['name'],
            row['location_names'] or 'N/A',
            row['rack_names'] or 'N/A',
            objtype_name,  # Usar nome do tipo em vez do ID
            row['asset_no'] or 'N/A'
        )
        processed_data.append(processed_row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    
    title = "Relatório de todos os equipamentos"
    logo_path = "logo-coids.png"
    headers = ["Nome", "Localizacoes", "Racks", "Tipo", "Asset No."]
    apply_excel_styles(ws, title, headers, processed_data, logo_path)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=equipamentos.xlsx"}
    )


# Export Lista de responsáveis
@exports_bp.route("/api/contacts/export/xlsx")
@db_connection
def export_all_contacts_xlsx(db):
    cursor = db.cursor()
    cursor.execute("""
        SELECT 
            DISTINCT av.string_value AS contact_name,
            COUNT(DISTINCT o.id) AS equipment_count
        FROM 
            AttributeValue av
        JOIN
            Attribute a ON av.attr_id = a.id
        JOIN
            Object o ON av.object_id = o.id
        WHERE
            a.name = 'contact person'
        GROUP BY 
            av.string_value
        ORDER BY 
            contact_name
    """)
    data = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responsaveis"
    
    title = "Relatório de Responsáveis"
    logo_path = "logo-coids.png"
    headers = ["Responsável", "Quantidade de Equipamentos"]
    apply_excel_styles(ws, title, headers, data, logo_path)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=responsaveis.xlsx"}
    )


# Export Lista de Equipamentos por Responsável 
@exports_bp.route("/api/contacts/<string:contact_name>/export/xlsx")
@db_connection
def export_contact_equipment_xlsx(db, contact_name):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT 
            o.id,
            o.name, 
            o.objtype_id,
            o.asset_no,
            GROUP_CONCAT(DISTINCT r.name) as rack_names,
            GROUP_CONCAT(DISTINCT l.name) as location_names,
            GROUP_CONCAT(DISTINCT rs.unit_no) as unit_nos
        FROM Object o
        JOIN AttributeValue av ON o.id = av.object_id
        JOIN Attribute a ON av.attr_id = a.id
        LEFT JOIN RackSpace rs ON o.id = rs.object_id
        LEFT JOIN Rack r ON rs.rack_id = r.id
        LEFT JOIN Location l ON r.location_id = l.id
        WHERE a.name = 'contact person' 
          AND av.string_value = %s
          AND o.objtype_id NOT IN (1560, 1561, 1562)
        GROUP BY o.id  -- ADICIONAR GROUP BY PARA EVITAR DUPLICATAS
        ORDER BY o.name
    """, (contact_name,))
    data = cursor.fetchall()

    # Processar dados para incluir objtype_name
    processed_data = []
    for row in data:
        objtype_name = get_object_type_name(row['objtype_id'])
        processed_row = (
            row['name'],
            objtype_name,
            row['asset_no'] or 'N/A',
            row['rack_names'] or 'N/A',
            row['location_names'] or 'N/A',
            row['unit_nos'] or 'N/A'
        )
        processed_data.append(processed_row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    
    title = f"Equipamentos do Responsável: {contact_name}"
    logo_path = "logo-coids.png"
    headers = ["Nome", "Tipo", "Asset No.", "Racks", "Localizações", "Unidades"]
    apply_excel_styles(ws, title, headers, processed_data, logo_path)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=equipamentos_{contact_name}.xlsx"}
    )



# Export Lista de Equipamentos Filtrados por Tipo
@exports_bp.route("/api/search/equipments/export/xlsx")
@db_connection
def export_equipments_filtered_xlsx(db):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    
    # Obter parâmetros de filtro
    objtype_id = request.args.get('type_id', '')
    
    # Construir query base
    query = """
        SELECT 
            o.id,
            o.name, 
            GROUP_CONCAT(DISTINCT l.name) as location_names,
            GROUP_CONCAT(DISTINCT r.name) as rack_names,
            o.objtype_id, 
            o.asset_no
        FROM Object o
        LEFT JOIN RackSpace rs ON o.id = rs.object_id
        LEFT JOIN Rack r ON rs.rack_id = r.id
        LEFT JOIN Location l ON r.location_id = l.id
        WHERE o.objtype_id NOT IN (1560, 1561, 1562)
    """
    
    params = []
    
    # Aplicar filtro de tipo se fornecido
    if objtype_id:
        query += " AND o.objtype_id = %s"
        params.append(objtype_id)
    
    query += " GROUP BY o.id ORDER BY o.name"
    
    cursor.execute(query, params)
    data = cursor.fetchall()

    # Processar dados para incluir objtype_name
    processed_data = []
    for row in data:
        objtype_name = get_object_type_name(row['objtype_id'])
        processed_row = (
            row['name'],
            row['location_names'] or 'N/A',
            row['rack_names'] or 'N/A',
            objtype_name,
            row['asset_no'] or 'N/A'
        )
        processed_data.append(processed_row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    
    # Definir título baseado no filtro
    if objtype_id:
        type_name = get_object_type_name(int(objtype_id))
        title = f"Relatório de Equipamentos - {type_name}"
        filename = f"equipamentos_{type_name.lower().replace(' ', '_')}.xlsx"
    else:
        title = "Relatório de todos os equipamentos"
        filename = "equipamentos.xlsx"
    
    logo_path = "logo-coids.png"
    headers = ["Nome", "Localizacoes", "Racks", "Tipo", "Asset No."]
    apply_excel_styles(ws, title, headers, processed_data, logo_path)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )


# Export PDF - Detalhes do Equipamento
@exports_bp.route("/api/object/<int:object_id>/export-pdf")
@db_connection
def export_object_pdf(db, object_id):
    try:
        from flask import send_file
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from datetime import datetime
        import html
        from io import BytesIO
        
        cursor = db.cursor(MySQLdb.cursors.DictCursor)
        
        # Buscar dados do objeto
        cursor.execute("""
            SELECT id, name, label, objtype_id, asset_no, has_problems, comment
            FROM Object 
            WHERE id = %s
        """, (object_id,))
        
        obj = cursor.fetchone()
        if not obj:
            cursor.close()
            return jsonify({"error": "Objeto não encontrado"}), 404
        
        from src.utils.object_types import get_object_type_name
        obj['objtype_name'] = get_object_type_name(obj['objtype_id'])
        
        # Informações do rack
        cursor.execute("""
            SELECT r.name as rack_name, rs.unit_no, l.name as location_name
            FROM RackSpace rs
            JOIN Rack r ON rs.rack_id = r.id
            LEFT JOIN Location l ON r.location_id = l.id
            WHERE rs.object_id = %s
        """, (object_id,))
        rack_info = cursor.fetchone()
        
        # ATRIBUTOS 
        cursor.execute("""
            SELECT a.name as attribute_name,
                   COALESCE(av.string_value, av.uint_value, av.float_value) as attribute_value
            FROM AttributeValue av 
            JOIN Attribute a ON av.attr_id = a.id 
            WHERE av.object_id = %s
        """, (object_id,))
        raw_attributes = cursor.fetchall()
        
        # PROCESSAR ATRIBUTOS
        processed_attributes = []
        
        # Carregar mapeamentos
        from src.routes.objects import get_hw_type_map, get_sw_type_map
        hw_type_map = get_hw_type_map(db)
        sw_type_map = get_sw_type_map(db)
        
        for attr in raw_attributes:
            attribute_name = attr['attribute_name']
            attribute_value = attr['attribute_value']
            
            # Format support contract expiration
            if attribute_name == 'support contract expiration' and attribute_value:
                try:
                    from datetime import datetime, timezone
                    dt = datetime.fromtimestamp(int(attribute_value), tz=timezone.utc)
                    attr['attribute_value'] = dt.strftime('%Y-%m-%d') 
                except:
                    pass
            
            # Format HW warranty expiration
            elif attribute_name == 'HW warranty expiration' and attribute_value:
                try:
                    from datetime import datetime, timezone
                    dt = datetime.fromtimestamp(int(attribute_value), tz=timezone.utc)
                    attr['attribute_value'] = dt.strftime('%Y-%m-%d') 
                except:
                    pass
            
            # Mapear HW type
            elif attribute_name == 'HW type' and attribute_value:
                mapped_value = hw_type_map.get(str(attribute_value), f"HW Type {attribute_value}")
                attr['attribute_value'] = mapped_value

            # Mapear SW type
            elif attribute_name == 'SW type' and attribute_value:
                mapped_value = sw_type_map.get(str(attribute_value), f"SW Type {attribute_value}")
                attr['attribute_value'] = mapped_value
            
            # Mapear Hypervisor
            elif attribute_name == 'Hypervisor' and attribute_value:
                hypervisor_map = {
                    '1501': 'Sim',
                    '1502': 'Não',
                }
                mapped_value = hypervisor_map.get(str(attribute_value), f"Hypervisor {attribute_value}")
                attr['attribute_value'] = mapped_value

            processed_attributes.append(attr)

        # Portas de rede
        ports = []
        try:
            cursor.execute("""
                SELECT
                    p.name AS port_name,
                    p.label AS port_label,
                    p.l2address AS l2_address,
                    COALESCE(oi.oif_name, ii.iif_name) AS interface_name
                FROM Port p
                LEFT JOIN PortOuterInterface oi ON oi.id = p.type
                LEFT JOIN PortInnerInterface ii ON ii.id = p.type
                WHERE p.object_id = %s
                ORDER BY p.name;
            """, (object_id,))

            raw_ports = cursor.fetchall()

            for port in raw_ports:
                processed_port = {
                    'port_name': port['port_name'],
                    'port_label': port['port_label'] or '',
                    'interface_name': port['interface_name'] or 'N/A',
                    'l2_address': port['l2_address'] or 'N/A'
                }
                ports.append(processed_port)

        except MySQLdb.Error:
            ports = []

        
        cursor.close()
        
        # Decodificar entidades HTML (usando função do objects.py)
        from src.routes.objects import decode_html_entities
        if obj.get('comment'):
            obj['comment'] = decode_html_entities(obj['comment'])
        
        # Criar PDF 
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilos
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor("#0077BA"),
            alignment=1,
            spaceAfter=20
        )
        
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.HexColor("#34495e")
        )
        
        # Título
        elements.append(Paragraph(f"FICHA TÉCNICA - {obj['name']}", title_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Informações Básicas
        elements.append(Paragraph("RESUMO", heading_style))
        
        basic_data = [
            ['Common name:', obj['name'] or 'N/A'],
            ['Object type:', obj['objtype_name'] or 'N/A'],
            ['Visible label:', obj['label'] or 'N/A'],
            ['Asset Tag:', obj['asset_no'] or 'N/A'],
            ['has_problems:', obj['has_problems'] or 'N/A'],
        ]
        
        if rack_info:
            basic_data.extend([
                ['Rack:', rack_info['rack_name'] or 'N/A'],
                ['rack position:', rack_info['unit_no'] or 'N/A'],
                ['location:', rack_info['location_name'] or 'N/A']
            ])
        
        basic_table = Table(basic_data, colWidths=[2*inch, 4*inch])
        basic_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(basic_table)
        
        # ATRIBUTOS
        if processed_attributes:
            elements.append(Paragraph("ATRIBUTOS", heading_style))
            
            attr_data = [['Atributo', 'Valor']]
            for attr in processed_attributes: # Usando processed_attributes
                value = str(attr['attribute_value'])[:100] + ('...' if len(str(attr['attribute_value'])) > 100 else '')
                attr_data.append([attr['attribute_name'], value])
            
            attr_table = Table(attr_data, colWidths=[2.5*inch, 3.5*inch])
            attr_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0077BA")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ]))
            elements.append(attr_table)
        
        # Portas de Rede
        if ports:
            elements.append(Paragraph("PORTAS DE REDE", heading_style))
            
            ports_data = [['Local name', 'Visible label', 'Interface', 'L2 address', 'Remote object and port', 'Cable ID']]
            for port in ports:
                ports_data.append([
                    port['port_name'],                    # Local name
                    port['port_label'],                   # Visible label  
                    port['interface_name'],               # Interface
                    port['l2_address'],                   # L2 address
                    '',                                   # Remote object and port
                    ''                                    # Cable ID
                ])
            
            ports_table = Table(ports_data, colWidths=[0.9*inch, 0.8*inch, 0.9*inch, 1*inch, 1.5*inch, 1*inch])
            ports_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0077BA")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ]))
            elements.append(ports_table)
        
        # Comentários
        if obj.get('comment'):
            elements.append(Paragraph("COMENTÁRIOS", heading_style))
            comment = obj['comment']  # Já foi decodificado
            if len(comment) > 500:
                comment = comment[:500] + "... [texto truncado]"
            comment_para = Paragraph(comment.replace('\n', '<br/>'), styles['Normal'])
            elements.append(comment_para)
        
        # Rodapé
        elements.append(Spacer(1, 0.3*inch))
        footer_style = ParagraphStyle('Footer', fontSize=8, textColor=colors.grey, alignment=1)
        elements.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", footer_style))
        
        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Nome do arquivo
        safe_name = "".join(c for c in obj['name'] if c.isalnum() or c in (' ', '-', '_')).rstrip()
        filename = f"ficha_tecnica_{safe_name}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Erro na geração do PDF: {e}")
        return jsonify({"error": f"Erro na geração do PDF: {str(e)}"}), 500



# Exportar PDF - Lista de equipamentos por rack
@exports_bp.route("/api/objects/rack/<int:rack_id>/export/pdf")
@db_connection
def export_rack_objects_pdf(db, rack_id):
    cursor = db.cursor(MySQLdb.cursors.DictCursor) 
    
    # Nome do rack
    cursor.execute("SELECT name FROM Rack WHERE id = %s", (rack_id,))
    rack_result = cursor.fetchone()
    rack_name = rack_result['name'] if rack_result else f"Rack_{rack_id}"
    
    # Dados dos equipamentos
    cursor.execute("""
        SELECT 
            o.id, o.name, o.objtype_id, o.asset_no,
            rs.unit_no
        FROM RackSpace rs
        JOIN Object o ON rs.object_id = o.id
        WHERE rs.rack_id = %s
        ORDER BY rs.unit_no DESC
    """, (rack_id,))
    data = cursor.fetchall()
    
    # Processar dados para incluir objtype_name
    processed_data = []
    for row in data:
        objtype_name = get_object_type_name(row['objtype_id'])
        processed_data.append([
            row['id'],
            row['name'],
            objtype_name,  # Usar nome do tipo
            row['asset_no'] or 'N/A',
            row['unit_no'] or 'N/A'
        ])
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Criar a tabela de dados
    table_data = [["ID", "Nome", "Tipo", "Asset No.", "Unidade"]]
    table_data.extend(processed_data)

    table = Table(table_data)
    
    # Estilo da tabela
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    
    story = [table]
    doc.build(story)
    
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment;filename={rack_name}_equipamentos.pdf"}
    )