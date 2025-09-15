from flask import Flask, jsonify, request, Response, Blueprint
from flask_cors import CORS
import MySQLdb
from functools import wraps
import csv
from io import StringIO, BytesIO
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


app = Flask(__name__)
CORS(app)  # Habilita CORS para o React

def get_db():
    return MySQLdb.connect(
        host=os.getenv('DB_HOST'),
        user=os.getenv('DB_USER'),
        passwd=os.getenv('DB_PASSWORD'),
        db=os.getenv('DB_NAME'),
        charset=os.getenv('DB_CHARSET', 'utf8'),
        port=int(os.getenv('DB_PORT', 3306))
    )
    
def db_connection(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        db = get_db()
        try:
            return f(db, *args, **kwargs)
        finally:
            db.close()
    return decorated_function

# ***********ROTAS DAS APIs***********************************************
# Rota para lista de racks
@app.route("/api/racks")
@db_connection
def get_racks(db):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("""
        SELECT 
            r.id, r.name, r.height, r.row_name,
            r.location_id, r.location_name,
            r.asset_no, r.comment,
            COUNT(DISTINCT rs.object_id) as object_count
        FROM Rack r
        LEFT JOIN RackSpace rs ON r.id = rs.rack_id
        GROUP BY r.id
        ORDER BY r.name
    """)
    
    data = cursor.fetchall()
    cursor.close()
    return jsonify(data)

# Rota para detalhes do rack
@app.route("/api/rack/<int:rack_id>")
@db_connection
def get_rack_detail(db, rack_id):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    
    # Informações do rack
    cursor.execute("""
        SELECT id, name, height, row_name, 
               location_id, location_name, asset_no, comment
        FROM Rack 
        WHERE id = %s
    """, (rack_id,))
    
    rack = cursor.fetchone()
    
    if not rack:
        cursor.close()
        return jsonify({"error": "Rack não encontrado"}), 404
    
    # Busca objetos no rack usando RackSpace
    cursor.execute("""
        SELECT 
            o.id, o.name, o.objtype_id, o.asset_no,
            o.has_problems, o.comment,
            rs.unit_no, rs.atom, rs.state
        FROM RackSpace rs
        JOIN Object o ON rs.object_id = o.id
        WHERE rs.rack_id = %s
        ORDER BY rs.unit_no DESC
    """, (rack_id,))
    
    objects = cursor.fetchall()
    cursor.close()
    
    return jsonify({
        'rack': rack,
        'objects': objects
    })

# *****************************************************************************

# Rota para objetos
@app.route("/api/objects")
@db_connection
def get_objects(db):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("""
        SELECT 
            o.id, o.name, o.objtype_id, o.asset_no,
            o.has_problems, o.comment,
            GROUP_CONCAT(DISTINCT r.name) as rack_names,
            GROUP_CONCAT(DISTINCT l.name) as location_names
        FROM Object o
        LEFT JOIN RackSpace rs ON o.id = rs.object_id
        LEFT JOIN Rack r ON rs.rack_id = r.id
        LEFT JOIN Location l ON r.location_id = l.id
        GROUP BY o.id
        ORDER BY o.name 
    """)
    
    data = cursor.fetchall()
    cursor.close()
    return jsonify(data)

@app.route("/api/objects/count")
@db_connection
def get_objects_count(db):
    cursor = db.cursor()
    
    cursor.execute("""
        SELECT COUNT(DISTINCT o.id) as unique_objects_count
        FROM Object o
    """)
    
    count = cursor.fetchone()[0]
    cursor.close()
    return jsonify({"unique_objects_count": count})


# Rota para detalhes do objeto
@app.route("/api/object/<int:object_id>")
@db_connection
def get_object_detail(db, object_id):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    
    # Informações básicas do objeto
    cursor.execute("""
        SELECT id, name, objtype_id, asset_no, 
               has_problems, comment
        FROM Object 
        WHERE id = %s
    """, (object_id,))
    
    obj = cursor.fetchone()
    
    if not obj:
        cursor.close()
        return jsonify({"error": "Objeto não encontrado"}), 404
    
    # Informações do rack
    cursor.execute("""
        SELECT 
            r.id as rack_id, 
            r.name as rack_name, 
            rs.unit_no,
            rs.atom,
            rs.state
        FROM RackSpace rs
        JOIN Rack r ON rs.rack_id = r.id
        WHERE rs.object_id = %s
    """, (object_id,))
    
    rack_info = cursor.fetchone()
    
    # Atributos do objeto
    cursor.execute("""
        SELECT 
    a.name as attribute_name, 
    a.type as attribute_type,
    COALESCE(av.string_value, av.uint_value, av.float_value) as attribute_value
FROM AttributeValue av 
JOIN Attribute a ON av.attr_id = a.id 
WHERE av.object_id = %s
    """, (object_id,))
    
    attributes = cursor.fetchall()
    
    # Portas de rede (se existirem)
    ports = []
    try:
        cursor.execute("""
            SELECT 
                name as port_name,
                type as port_type,
                label as port_label
            FROM Port 
            WHERE object_id = %s
            ORDER BY name
        """, (object_id,))
        ports = cursor.fetchall()
    except MySQLdb.Error:
        pass
    
    cursor.close()
    
    return jsonify({
        'object': obj,
        'rack': rack_info,
        'attributes': attributes,
        'ports': ports
    })

# Rota para buscar objetos por tipo
@app.route("/api/objects/type/<int:objtype_id>")
@db_connection
def get_objects_by_type(db, objtype_id):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("""
        SELECT 
            o.id, o.name, o.objtype_id, o.asset_no,
            r.name as rack_name,
            rs.unit_no as rack_unit,
            l.name as location_name
        FROM Object o
        LEFT JOIN RackSpace rs ON o.id = rs.object_id
        LEFT JOIN Rack r ON rs.rack_id = r.id
        LEFT JOIN Location l ON r.location_id = l.id
        WHERE o.objtype_id = %s
        ORDER BY o.name
    """, (objtype_id,))
    
    data = cursor.fetchall()
    cursor.close()
    return jsonify(data)

# *****************************************************************************

# Rota para buscar estatísticas
@app.route("/api/stats")
@db_connection
def get_stats(db):
    cursor = db.cursor(MySQLdb.cursors.DictCursor)
    
    stats = {}
    
    # Total de objetos
    cursor.execute("SELECT COUNT(*) as count FROM Object")
    stats['total_objects'] = cursor.fetchone()['count']
    
    # Total de racks
    cursor.execute("SELECT COUNT(*) as count FROM Rack")
    stats['total_racks'] = cursor.fetchone()['count']
    
    # Objetos por tipo
    cursor.execute("""
        SELECT objtype_id, COUNT(*) as count 
        FROM Object 
        GROUP BY objtype_id 
        ORDER BY count DESC
    """)
    stats['objects_by_type'] = cursor.fetchall()
    
    # Racks com mais objetos
    cursor.execute("""
        SELECT r.name, COUNT(rs.object_id) as object_count
        FROM Rack r
        LEFT JOIN RackSpace rs ON r.id = rs.rack_id
        GROUP BY r.id
        ORDER BY object_count DESC
        LIMIT 10
    """)
    stats['top_racks'] = cursor.fetchall()
    
    cursor.close()
    return jsonify(stats)

# ************************ROTAS PARA EXPORTAÇÃO*************************************
#************************ Rota para exportar xlsx ***************

# Função para estilizar a tabela
def apply_excel_styles(ws, title_text, headers, logo_path=None):
    logo_height_px = 70
    logo_width_px = 60
    
    # Inserir a logo
    if logo_path:
        img = Image(logo_path)
        img.height = logo_height_px
        img.width = logo_width_px
        ws.add_image(img, 'A1')

    # Definir a altura da linha 1 para acomodar a logo
    ws.row_dimensions[1].height = logo_height_px * 0.75 # Conversão de pixels para pontos

    # Título principal dinâmico
    num_cols = len(headers)
    end_col_letter = get_column_letter(num_cols)
    ws.merge_cells(f'B1:{end_col_letter}1')
    title_cell = ws['B1']
    title_cell.value = title_text
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Espaçamento
    #ws.append([])
    
    # Cabeçalho da tabela
    ws.append(headers)
    
    header_fill = PatternFill(start_color="0077BA", end_color="0077BA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    header_row_index = ws.max_row
    for cell in ws[header_row_index]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    # 6. Ajustar largura das colunas
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = len(header) + 5


# ************* Lista de RAcks ************************
@app.route("/api/racks/export/xlsx")
@db_connection
def export_racks_xlsx(db):
    cursor = db.cursor()
    cursor.execute("""
        SELECT 
            r.name, r.location_name, r.row_name, 
            r.height,
            COUNT(DISTINCT rs.object_id) as object_count
        FROM Rack r
        LEFT JOIN RackSpace rs ON r.id = rs.rack_id
        GROUP BY r.id
        ORDER BY r.name
    """)
    data = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Racks"

    title = "Relatório de Racks"
    logo_path = "logo-coids.png" 
    headers = ["Rack", "Localizacao", "Linha", "Altura", "Equipamentos"]
    apply_excel_styles(ws, title, headers, logo_path)

    for row in data:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=racks.xlsx"}
    )

#*************** Lista de Equipamentos *******************
@app.route("/api/objects/export/xlsx")
@db_connection
def export_all_objects_xlsx(db):
    cursor = db.cursor()
    cursor.execute("""
        SELECT 
            o.id, o.name, o.objtype_id, o.asset_no,
            GROUP_CONCAT(DISTINCT r.name) as rack_names,
            GROUP_CONCAT(DISTINCT l.name) as location_names
        FROM Object o
        LEFT JOIN RackSpace rs ON o.id = rs.object_id
        LEFT JOIN Rack r ON rs.rack_id = r.id
        LEFT JOIN Location l ON r.location_id = l.id
        GROUP BY o.id
        ORDER BY o.name
    """)
    data = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    
    headers = ["ID", "Nome", "Tipo", "Asset No.", "Racks", "Localizacoes"]
    ws.append(headers)
    
    for row in data:
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=equipamentos.xlsx"}
    )

#************* Rota para exportar pdf ***************
#************* Lista de equipamentos ****************

@app.route("/api/objects/rack/<int:rack_id>/export/pdf")
@db_connection
def export_rack_objects_pdf(db, rack_id):
    cursor = db.cursor()
    
    # Nome do rack
    cursor.execute("SELECT name FROM Rack WHERE id = %s", (rack_id,))
    rack_name = cursor.fetchone()[0]
    
    # Dados dos equipamentos
    cursor.execute("""
        SELECT 
            o.id, o.name, o.objtype_id, o.asset_no,
            rs.unit_no
        FROM RackSpace rs
        JOIN Object o ON rs.object_id = o.id
        WHERE rs.rack_id = %s
        ORDER BY rs.unit_no DESC
    """, (rack_id,))
    data = cursor.fetchall()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Criar a tabela de dados
    table_data = [["ID", "Nome", "Tipo", "Asset No.", "Unidade"]]
    for row in data:
        table_data.append(list(row))

    table = Table(table_data)
    
    # Estilo da tabela
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    
    story = [table]
    doc.build(story)
    
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment;filename={rack_name}_equipamentos.pdf"}
    )


# *****************************************************************************

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)